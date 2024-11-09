from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

class ExcelWriter:
    def __init__(self, filename: str):
        self.filename = filename
        self.wb = Workbook()
        self.ws = self.wb.active
        self._setup_headers()
        
    def _setup_headers(self):
        """Excel fayl ustunlarini sozlash"""
        # Ustun kengliklari
        self.ws.column_dimensions['A'].width = 40  # Sarlavha
        self.ws.column_dimensions['B'].width = 50  # Tavsif
        self.ws.column_dimensions['C'].width = 15  # Ko'rishlar
        self.ws.column_dimensions['D'].width = 20  # Nashr vaqti
        
        # Sarlavhalar
        headers = ['Sarlavha', 'Tavsif', "Ko'rishlar soni", 'Nashr vaqti']
        header_font = Font(bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
    
    def write_articles(self, articles):
        """Yangiliklarni Excel fayliga yozish"""
        for row, article in enumerate(articles, 2):  # 2-qatordan boshlaymiz (1-qator sarlavha)
            # Sarlavha
            self.ws.cell(row=row, column=1, value=article.title)
            
            # Tavsif
            self.ws.cell(row=row, column=2, value=article.description)
            
            # Ko'rishlar soni
            views_cell = self.ws.cell(row=row, column=3)
            views_cell.value = article.views_count
            views_cell.alignment = Alignment(horizontal='center')
            
            # Nashr vaqti
            date_cell = self.ws.cell(row=row, column=4)
            date_cell.value = article.published_date
            date_cell.alignment = Alignment(horizontal='center')
            
            for col in range(1, 5):
                cell = self.ws.cell(row=row, column=col)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    def save(self):
        """Excel faylni saqlash"""
        try:
            self.wb.save(self.filename)
        except Exception as e:
            print(f"Excel faylni saqlashda xatolik: {str(e)}")